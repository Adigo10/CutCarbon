"""
Data export router — download raw or processed data as Excel (.xlsx), JSON, CSV, or PDF.
"""
import asyncio
import csv
import io
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.database import AgentRunDB, OffsetPurchaseDB, ScenarioDB, UserDB, get_db
from app.models.schemas import (
    OffsetPortfolioSummary,
    ScenarioComplianceOverrides,
    ScenarioReportMetric,
    ScenarioReportPayload,
)
from app.routers.auth import get_current_user
from app.services.financial_engine import get_compliance_report
from app.services.scenario_serializer import serialize_scenario
from app.utils.time import utcnow

router = APIRouter()

_DATA_DIR = Path(__file__).parent.parent / "data"

# CSV/Excel formula-injection mitigation. A string cell beginning with any of these
# is interpreted as a formula (or DDE payload) by Excel/LibreOffice/Sheets when the
# file is opened — dangerous because these reports are built for third-party auditors.
# Prefixing a single quote forces the cell to be treated as literal text.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value):
    """Neutralize spreadsheet formula injection for a single cell value."""
    if isinstance(value, str) and value[:1] in _FORMULA_PREFIXES:
        return "'" + value
    return value


class _SafeCsvWriter:
    """csv.writer wrapper that sanitizes every cell against formula injection."""

    def __init__(self, buf):
        self._w = csv.writer(buf)

    def writerow(self, row):
        self._w.writerow([_safe_cell(c) for c in row])

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)


def _pdf_text(value) -> str:
    """Escape user-derived text before it is parsed as reportlab Paragraph markup."""
    from xml.sax.saxutils import escape

    return escape(str(value))
_REPORT_METHODODOLOGY = "GHG Protocol Corporate Standard, ISO 14064-1"
_REPORT_DISCLAIMER = (
    "Calculations are based on industry-standard emission factors and scenario assumptions. "
    "Actual emissions may vary. For verified reporting, engage an accredited third-party verifier."
)
_CATEGORY_LABELS = [
    ("travel_tco2e", "Travel"),
    ("venue_energy_tco2e", "Venue Energy"),
    ("accommodation_tco2e", "Accommodation"),
    ("catering_tco2e", "Catering"),
    ("materials_waste_tco2e", "Materials & Waste"),
    ("equipment_tco2e", "Equipment"),
    ("swag_tco2e", "Swag"),
    ("digital_tco2e", "Digital & Virtual"),
]


def _wb_bytes(wb) -> bytes:
    """Sanitize + serialize a workbook (CPU-bound — call via asyncio.to_thread).

    Single chokepoint for every .xlsx export: neutralize formula injection in any
    string cell before serializing. openpyxl auto-types a leading "=" as a formula,
    so we rewrite dangerous string cells to literal text here.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value[:1] in _FORMULA_PREFIXES:
                    cell.value = "'" + cell.value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_response(content: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header(ws, row=1):
    """Bold + green background for header row."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1A9E6E")
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _labelize(value: str) -> str:
    return value.replace("_", " ").replace("/", " / ").title()


def _scenario_location(s: ScenarioDB) -> str:
    return getattr(s, "location", None) or (s.input_payload or {}).get("location", "")


async def _get_scenario_or_404(
    scenario_id: str,
    db: AsyncSession,
    current_user: UserDB,
) -> ScenarioDB:
    scenario = await db.scalar(
        select(ScenarioDB).where(
            ScenarioDB.id == scenario_id,
            ScenarioDB.user_id == current_user.id,
        )
    )
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
    return scenario


def _build_category_rows(scenario_data: dict[str, Any]) -> list[ScenarioReportMetric]:
    emissions = scenario_data["emissions"]
    total = emissions.get("total_tco2e") or 0
    rows: list[ScenarioReportMetric] = []
    for key, label in _CATEGORY_LABELS:
        value = float(emissions.get(key, 0) or 0)
        if value <= 0:
            continue
        pct_total = round((value / total) * 100, 1) if total > 0 else None
        rows.append(
            ScenarioReportMetric(
                key=key,
                label=label,
                value=round(value, 4),
                unit="tCO2e",
                pct_total=pct_total,
            )
        )
    return rows


# Modes counted as NZCE "Local Transportation" (venue-area movement) rather than
# "Travel to/from the Destination".
_NZCE_LOCAL_MODES = {"mrt_metro", "taxi_rideshare", "shuttle_bus", "bus_coach", "e_scooter", "cycling"}

_NZCE_BASE_NOTE = (
    "Indicative mapping to the Net Zero Carbon Events (NZCE) Measurement Methodology "
    "v1 categories, derived from this report's category totals."
)


def _build_nzce_rows(scenario_data: dict[str, Any]) -> tuple[list[ScenarioReportMetric], str]:
    """Map stored category totals onto the 9 NZCE categories.

    Travel is split into to/from-destination vs local transportation by re-running
    the per-segment factor math on the stored input payload; equipment is split into
    freight vs production/materials via its freight_tonne_km line.
    """
    from app.services.emissions_engine import EF

    emissions = scenario_data["emissions"]
    payload = scenario_data.get("input_payload") or {}
    total = float(emissions.get("total_tco2e") or 0)
    note = _NZCE_BASE_NOTE

    travel_total = float(emissions.get("travel_tco2e") or 0)
    local_travel = 0.0
    segments = payload.get("travel_segments") or []
    if segments and travel_total > 0:
        local_kg = 0.0
        all_kg = 0.0
        for seg in segments:
            ef_data = EF["travel"].get(seg.get("mode")) or {}
            if "economy" in ef_data:
                cls = seg.get("travel_class") or "economy"
                ef = ef_data.get(cls) or ef_data.get("business") or ef_data.get("economy") or 0
            else:
                ef = ef_data.get("factor", 0)
            kg = (seg.get("attendees") or 0) * (seg.get("distance_km") or 0) * ef
            all_kg += kg
            if seg.get("mode") in _NZCE_LOCAL_MODES:
                local_kg += kg
        if all_kg > 0:
            local_travel = round(travel_total * (local_kg / all_kg), 4)
    elif travel_total > 0:
        note += " Proxy travel is mapped entirely to Travel to/from the Destination."
    dest_travel = round(travel_total - local_travel, 4)

    equipment_total = float(emissions.get("equipment_tco2e") or 0)
    freight = 0.0
    freight_tkm = ((payload.get("equipment") or {}).get("freight_tonne_km")) or 0
    if freight_tkm and equipment_total > 0:
        freight_ef = EF.get("equipment", {}).get("freight_truck_per_km", {}).get("factor", 0.107)
        freight = round(min(equipment_total, freight_tkm * freight_ef / 1000), 4)
    production = round(float(emissions.get("swag_tco2e") or 0) + max(0.0, equipment_total - freight), 4)

    rows_spec = [
        ("nzce_production_materials", "Production & Materials", production),
        ("nzce_freight_logistics", "Freight & Logistics", freight),
        ("nzce_food_beverage", "Food & Beverage", emissions.get("catering_tco2e")),
        ("nzce_travel_destination", "Travel to/from the Destination", dest_travel),
        ("nzce_local_transportation", "Local Transportation", local_travel),
        ("nzce_accommodation", "Accommodation", emissions.get("accommodation_tco2e")),
        ("nzce_energy", "Energy", emissions.get("venue_energy_tco2e")),
        ("nzce_waste", "Waste", emissions.get("materials_waste_tco2e")),
        ("nzce_digital", "Digital Content & Communication", emissions.get("digital_tco2e")),
    ]
    rows = [
        ScenarioReportMetric(
            key=key,
            label=label,
            value=round(float(value or 0), 4),
            unit="tCO2e",
            pct_total=round(float(value or 0) / total * 100, 1) if total > 0 else None,
        )
        for key, label, value in rows_spec
    ]
    return rows, note


async def _build_offset_summary_for_scenario(
    scenario_id: str,
    user_id: int,
    total_tco2e: float,
    db: AsyncSession,
) -> OffsetPortfolioSummary:
    result = await db.execute(
        select(OffsetPurchaseDB).where(
            OffsetPurchaseDB.user_id == user_id,
            OffsetPurchaseDB.scenario_id == scenario_id,
            OffsetPurchaseDB.status != "cancelled",
        )
    )
    purchases = result.scalars().all()

    total_purchased = sum(p.quantity_tco2e for p in purchases)
    total_retired = sum(p.quantity_tco2e for p in purchases if p.status == "retired")
    total_cost = sum(p.total_cost_usd for p in purchases)

    by_type: dict[str, float] = {}
    by_registry: dict[str, float] = {}
    for purchase in purchases:
        by_type[purchase.project_type] = by_type.get(purchase.project_type, 0) + purchase.quantity_tco2e
        by_registry[purchase.registry] = by_registry.get(purchase.registry, 0) + purchase.quantity_tco2e

    coverage_pct = None
    if total_tco2e > 0:
        coverage_pct = round(total_retired / total_tco2e * 100, 1)

    return OffsetPortfolioSummary(
        total_purchased_tco2e=round(total_purchased, 3),
        total_retired_tco2e=round(total_retired, 3),
        total_cost_usd=round(total_cost, 2),
        by_project_type=by_type,
        by_registry=by_registry,
        coverage_pct=coverage_pct,
    )


async def build_scenario_report_payload(
    scenario_id: str,
    db: AsyncSession,
    current_user: UserDB,
    region: str = "singapore",
    has_scope3: bool = True,
    has_ghg_report: bool = False,
) -> ScenarioReportPayload:
    scenario = await _get_scenario_or_404(scenario_id, db, current_user)
    scenario_data = serialize_scenario(scenario)
    categories = _build_category_rows(scenario_data)
    nzce_categories, nzce_note = _build_nzce_rows(scenario_data)
    emissions = scenario_data["emissions"]
    offset_summary = await _build_offset_summary_for_scenario(
        scenario_id=scenario.id,
        user_id=current_user.id,
        total_tco2e=float(emissions.get("total_tco2e") or 0),
        db=db,
    )
    compliance = get_compliance_report(
        total_tco2e=float(emissions.get("total_tco2e") or 0),
        has_scope3=has_scope3,
        has_ghg_report=has_ghg_report,
        region=region,
        event_days=int(scenario.event_days or 0),
        attendees=int(scenario.attendees or 0),
    )

    return ScenarioReportPayload(
        report_title=f"Carbon Footprint Report - {scenario.event_name or scenario.name}",
        exported_at=utcnow().isoformat(),
        methodology=_REPORT_METHODODOLOGY,
        disclaimer=_REPORT_DISCLAIMER,
        scenario=scenario_data,
        categories=categories,
        scope_breakdown=emissions["scopes"],
        benchmark=scenario_data.get("benchmark"),
        assumptions=scenario_data.get("assumptions", {}),
        factor_snapshot=scenario_data.get("factors_snapshot", {}),
        offset_portfolio=offset_summary,
        compliance=compliance,
        compliance_overrides=ScenarioComplianceOverrides(
            region=region,
            has_scope3=has_scope3,
            has_ghg_report=has_ghg_report,
        ),
        nzce_categories=nzce_categories,
        nzce_note=nzce_note,
    )


def _scenario_report_filename(prefix: str, scenario_id: str, extension: str) -> str:
    stamp = utcnow().strftime("%Y%m%d")
    return f"{prefix}_{scenario_id}_{stamp}.{extension}"


def _csv_response(content: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pdf_response(content: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([content]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _scenario_report_csv_bytes(report: ScenarioReportPayload) -> bytes:
    buf = io.StringIO(newline="")
    writer = _SafeCsvWriter(buf)
    writer.writerow(["section", "key", "label", "value", "unit"])

    scenario = report.scenario
    emissions = scenario["emissions"]
    metadata_rows = [
        ("metadata", "scenario_id", "Scenario ID", scenario["scenario_id"], ""),
        ("metadata", "name", "Scenario Name", scenario["name"], ""),
        ("metadata", "event_name", "Event Name", scenario["event_name"], ""),
        ("metadata", "location", "Location", scenario["location"], ""),
        ("metadata", "event_type", "Event Type", scenario["event_type"], ""),
        ("metadata", "attendees", "Attendees", scenario["attendees"], "count"),
        ("metadata", "event_days", "Event Days", scenario["event_days"], "days"),
        ("metadata", "mode", "Mode", scenario.get("mode", ""), ""),
        ("metadata", "data_quality", "Data Quality", emissions["data_quality"], ""),
        ("metadata", "created_at", "Created At", scenario["created_at"], ""),
        ("metadata", "exported_at", "Exported At", report.exported_at, ""),
        ("metadata", "region", "Compliance Region", report.compliance_overrides.region, ""),
        ("metadata", "has_scope3", "Has Scope 3", report.compliance_overrides.has_scope3, "boolean"),
        ("metadata", "has_ghg_report", "Has GHG Report", report.compliance_overrides.has_ghg_report, "boolean"),
    ]
    writer.writerows(metadata_rows)

    for category in report.categories:
        writer.writerow(("categories", category.key, category.label, category.value, category.unit))
        if category.pct_total is not None:
            writer.writerow(
                ("categories_pct", f"{category.key}_pct_total", f"{category.label} % of total", category.pct_total, "pct")
            )

    for row in report.nzce_categories:
        writer.writerow(("nzce", row.key, row.label, row.value, row.unit))
    if report.nzce_note:
        writer.writerow(("nzce", "note", "NZCE Mapping Note", report.nzce_note, ""))

    for key, value in report.scope_breakdown.model_dump().items():
        writer.writerow(("scopes", key, _labelize(key), value, "tCO2e"))

    benchmark = report.benchmark.model_dump() if report.benchmark else {}
    for key, value in benchmark.items():
        unit = "tCO2e" if isinstance(value, (int, float)) else ""
        writer.writerow(("benchmark", key, _labelize(key), value, unit))

    for key, value in report.offset_portfolio.model_dump().items():
        if isinstance(value, dict):
            for subkey, subvalue in value.items():
                writer.writerow(("offsets", f"{key}.{subkey}", _labelize(f"{key} {subkey}"), subvalue, "tCO2e"))
        else:
            unit = "usd" if "cost" in key else ("pct" if key.endswith("_pct") else "tCO2e")
            writer.writerow(("offsets", key, _labelize(key), value, unit))

    writer.writerow(("compliance", "overall_score_pct", "Overall Score", report.compliance.overall_score_pct, "pct"))
    writer.writerow(("compliance", "penalty_risk_usd", "Penalty Risk", report.compliance.penalty_risk_usd, "usd"))
    for framework in report.compliance.mandatory_frameworks:
        writer.writerow(("compliance_mandatory", framework, framework, framework, ""))
    for index, check in enumerate(report.compliance.checks, start=1):
        prefix = f"check_{index}"
        writer.writerow(("compliance_check", f"{prefix}.framework", "Framework", check.framework, ""))
        writer.writerow(("compliance_check", f"{prefix}.status", f"{check.framework} status", check.status, ""))
        writer.writerow(("compliance_check", f"{prefix}.score_pct", f"{check.framework} score", check.score_pct, "pct"))
        for gap_idx, gap in enumerate(check.gaps, start=1):
            if gap:
                writer.writerow(("compliance_gap", f"{prefix}.gap_{gap_idx}", f"{check.framework} gap", gap, ""))
        for rec_idx, rec in enumerate(check.recommendations, start=1):
            writer.writerow(("compliance_recommendation", f"{prefix}.recommendation_{rec_idx}", f"{check.framework} recommendation", rec, ""))

    for key, value in report.assumptions.items():
        if isinstance(value, dict):
            for subkey, subvalue in value.items():
                writer.writerow(("assumptions", f"{key}.{subkey}", _labelize(f"{key} {subkey}"), subvalue, ""))
        else:
            writer.writerow(("assumptions", key, _labelize(key), value, ""))

    for key, value in report.factor_snapshot.items():
        unit = ""
        if isinstance(value, (int, float)):
            if "usd" in key:
                unit = "usd"
            elif "pct" in key:
                unit = "pct"
        writer.writerow(("factors", key, _labelize(key), value, unit))

    return buf.getvalue().encode("utf-8")


def _scenario_report_xlsx(report: ScenarioReportPayload):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Report Summary"
    summary_ws.append(["Field", "Value"])
    _style_header(summary_ws)

    scenario = report.scenario
    emissions = scenario["emissions"]
    for label, value in [
        ("Scenario", scenario["name"]),
        ("Event", scenario["event_name"]),
        ("Location", scenario["location"]),
        ("Event Type", scenario["event_type"]),
        ("Attendees", scenario["attendees"]),
        ("Event Days", scenario["event_days"]),
        ("Mode", scenario.get("mode", "")),
        ("Exported At", report.exported_at),
        ("Compliance Region", report.compliance_overrides.region),
        ("Has Scope 3", report.compliance_overrides.has_scope3),
        ("Has GHG Report", report.compliance_overrides.has_ghg_report),
        ("Total tCO2e", emissions["total_tco2e"]),
        ("Per Attendee tCO2e", emissions["per_attendee_tco2e"]),
        ("Per Attendee Day tCO2e", emissions["per_attendee_day_tco2e"]),
        ("Data Quality", emissions["data_quality"]),
        ("Overall Compliance Score", report.compliance.overall_score_pct),
        ("Offset Coverage %", report.offset_portfolio.coverage_pct if report.offset_portfolio.coverage_pct is not None else "—"),
    ]:
        summary_ws.append([label, value])

    summary_ws.append([])
    summary_ws.append(["Category", "tCO2e", "% of Total"])
    _style_header(summary_ws, summary_ws.max_row)
    categories_start = summary_ws.max_row + 1
    for category in report.categories:
        summary_ws.append([category.label, category.value, category.pct_total if category.pct_total is not None else ""])

    if report.categories:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Emissions by Category"
        chart.y_axis.title = "tCO2e"
        data = Reference(summary_ws, min_col=2, min_row=categories_start, max_row=summary_ws.max_row)
        labels = Reference(summary_ws, min_col=1, min_row=categories_start, max_row=summary_ws.max_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        chart.width = 12
        chart.height = 8
        summary_ws.add_chart(chart, "E4")

    nzce_ws = wb.create_sheet("NZCE Mapping")
    nzce_ws.append(["NZCE Category", "tCO2e", "% of Total"])
    _style_header(nzce_ws)
    for row in report.nzce_categories:
        nzce_ws.append([row.label, row.value, row.pct_total if row.pct_total is not None else ""])
    if report.nzce_note:
        nzce_ws.append([])
        nzce_ws.append(["Note", report.nzce_note])

    compliance_ws = wb.create_sheet("Compliance")
    compliance_ws.append(["Framework", "Status", "Score %", "Gaps", "Recommendations"])
    _style_header(compliance_ws)
    for check in report.compliance.checks:
        compliance_ws.append([
            check.framework,
            check.status,
            check.score_pct,
            "\n".join([gap for gap in check.gaps if gap]),
            "\n".join(check.recommendations),
        ])

    offsets_ws = wb.create_sheet("Offsets")
    offsets_ws.append(["Metric", "Value"])
    _style_header(offsets_ws)
    for key, value in report.offset_portfolio.model_dump().items():
        if isinstance(value, dict):
            for subkey, subvalue in value.items():
                offsets_ws.append([_labelize(f"{key} {subkey}"), subvalue])
        else:
            offsets_ws.append([_labelize(key), value])

    factors_ws = wb.create_sheet("Factors")
    factors_ws.append(["Key", "Value"])
    _style_header(factors_ws)
    for key, value in report.factor_snapshot.items():
        factors_ws.append([key, value])

    assumptions_ws = wb.create_sheet("Assumptions")
    assumptions_ws.append(["Key", "Value"])
    _style_header(assumptions_ws)
    if report.assumptions:
        for key, value in report.assumptions.items():
            if isinstance(value, dict):
                for subkey, subvalue in value.items():
                    assumptions_ws.append([f"{key}.{subkey}", str(subvalue)])
            else:
                assumptions_ws.append([key, value])
    else:
        assumptions_ws.append(["notes", "No assumptions recorded"])

    for ws in wb.worksheets:
        for column in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in column), default=0)
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 48)
        if ws.title == "Report Summary":
            ws["A1"].font = Font(bold=True)

    return wb


def _scenario_report_pdf(report: ScenarioReportPayload) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.platypus import (
        KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    page_width, page_height = A4
    brand_dark = colors.HexColor("#115e59")
    brand_green = colors.HexColor("#22c55e")
    ink = colors.HexColor("#111827")
    muted = colors.HexColor("#6b7280")
    copyright_year = (report.exported_at or "")[:4] or "2026"

    def _draw_logo(c, x, y, size):
        # Native redraw of the brand mark (frontend/public/favicon.svg): teal->green
        # gradient rounded square with a white "C". No SVG dependency needed.
        c.saveState()
        clip = c.beginPath()
        clip.roundRect(x, y, size, size, size * 0.25)
        c.clipPath(clip, stroke=0, fill=0)
        c.linearGradient(x, y + size, x + size, y, (brand_dark, brand_green))
        c.restoreState()
        c.saveState()
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", size * 0.62)
        c.drawCentredString(x + size / 2, y + size * 0.26, "C")
        c.restoreState()

    def _decorate_page(c, page_num, total_pages):
        """Branded header + copyright footer drawn on every page."""
        c.saveState()
        left = 1.5 * cm
        right = page_width - 1.5 * cm

        if page_num == 1:
            logo_size = 1.0 * cm
            logo_y = page_height - 1.9 * cm
            _draw_logo(c, left, logo_y, logo_size)
            c.setFillColor(ink)
            c.setFont("Helvetica-Bold", 13)
            c.drawString(left + logo_size + 0.35 * cm, logo_y + 0.52 * cm, "CutCarbon")
            c.setFillColor(muted)
            c.setFont("Helvetica", 8.5)
            c.drawString(
                left + logo_size + 0.35 * cm, logo_y + 0.12 * cm,
                "EventCarbon Co-Pilot — Event Carbon Footprint Report",
            )
            rule_y = page_height - 2.15 * cm
        else:
            logo_size = 0.55 * cm
            logo_y = page_height - 1.35 * cm
            _draw_logo(c, left, logo_y, logo_size)
            c.setFillColor(muted)
            c.setFont("Helvetica", 8)
            c.drawString(left + logo_size + 0.25 * cm, logo_y + 0.16 * cm, report.report_title)
            rule_y = page_height - 1.55 * cm

        c.setStrokeColor(colors.HexColor("#1A9E6E"))
        c.setLineWidth(1)
        c.line(left, rule_y, right, rule_y)

        c.setLineWidth(0.8)
        c.line(left, 1.55 * cm, right, 1.55 * cm)
        c.setFillColor(muted)
        c.setFont("Helvetica", 7.5)
        c.drawString(left, 1.15 * cm, f"© {copyright_year} CutCarbon · EventCarbon Co-Pilot. All rights reserved.")
        c.drawString(
            left, 0.78 * cm,
            "Confidential — prepared for the event organizer and its auditors. Not a third-party verification statement.",
        )
        c.drawRightString(right, 1.15 * cm, f"Page {page_num} of {total_pages}")
        c.restoreState()

    class _ReportCanvas(pdfcanvas.Canvas):
        """Buffers pages so the footer can state 'Page X of Y'."""

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._page_states = []

        def showPage(self):
            self._page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._page_states)
            for index, state in enumerate(self._page_states, start=1):
                self.__dict__.update(state)
                _decorate_page(self, index, total)
                super().showPage()
            super().save()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2.1 * cm,
        title=report.report_title,
        author="CutCarbon EventCarbon Co-Pilot",
        subject="Event carbon footprint report",
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], textColor=colors.HexColor("#1A9E6E")))
    styles.add(ParagraphStyle(name="SectionHeading", parent=styles["Heading2"], textColor=colors.HexColor("#1f6f6d")))
    styles.add(ParagraphStyle(name="BodySmall", parent=styles["BodyText"], fontSize=9, leading=12))

    def _table(rows: list[list[Any]], col_widths: Optional[list[float]] = None) -> Table:
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A9E6E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        return table

    scenario = report.scenario
    emissions = scenario["emissions"]
    benchmark = report.benchmark.model_dump() if report.benchmark else None
    offset_summary = report.offset_portfolio.model_dump()
    factor_rows = [["Factor", "Value"]]
    for key, value in report.factor_snapshot.items():
        factor_rows.append([_labelize(key), str(value)])

    assumption_paragraphs = []
    for key, value in report.assumptions.items():
        if isinstance(value, dict):
            value = ", ".join(f"{k}: {v}" for k, v in value.items())
        assumption_paragraphs.append(
            Paragraph(f"<b>{_pdf_text(_labelize(key))}:</b> {_pdf_text(value)}", styles["BodySmall"])
        )

    story = [
        Paragraph(report.report_title, styles["ReportTitle"]),
        Spacer(1, 0.35 * cm),
        Paragraph(f"Generated {report.exported_at}", styles["BodyText"]),
        Spacer(1, 0.2 * cm),
        Paragraph(
            f"{_pdf_text(scenario['name'])} in {_pdf_text(scenario['location'])} covers {scenario['attendees']} attendees across "
            f"{scenario['event_days']} day(s). Total modeled emissions are {emissions['total_tco2e']:.3f} tCO2e.",
            styles["BodyText"],
        ),
        Spacer(1, 0.35 * cm),
        _table(
            [
                ["Field", "Value"],
                ["Scenario", scenario["name"]],
                ["Event", scenario["event_name"]],
                ["Event Type", _labelize(scenario["event_type"])],
                ["Attendees", scenario["attendees"]],
                ["Event Days", scenario["event_days"]],
                ["Data Quality", emissions["data_quality"]],
                ["Compliance Region", report.compliance_overrides.region],
            ],
            [5 * cm, 11 * cm],
        ),
        Spacer(1, 0.35 * cm),
        Paragraph(f"<b>Methodology:</b> {report.methodology}", styles["BodySmall"]),
        Spacer(1, 0.15 * cm),
        Paragraph(f"<b>Disclaimer:</b> {report.disclaimer}", styles["BodySmall"]),
        PageBreak(),
        Paragraph("Emissions by Category", styles["SectionHeading"]),
        Spacer(1, 0.2 * cm),
        _table(
            [["Category", "tCO2e", "% of Total"]]
            + [
                [category.label, f"{category.value:.4f}", f"{category.pct_total:.1f}%" if category.pct_total is not None else "—"]
                for category in report.categories
            ],
            [7 * cm, 3 * cm, 3 * cm],
        ),
        Spacer(1, 0.35 * cm),
        Paragraph("Scope Breakdown", styles["SectionHeading"]),
        Spacer(1, 0.2 * cm),
        _table(
            [["Scope", "tCO2e"]]
            + [[_labelize(key), f"{value:.4f}"] for key, value in report.scope_breakdown.model_dump().items()],
            [8 * cm, 3 * cm],
        ),
        Spacer(1, 0.35 * cm),
    ]

    if report.nzce_categories:
        story.extend(
            [
                Paragraph("NZCE Category Mapping", styles["SectionHeading"]),
                Spacer(1, 0.2 * cm),
                _table(
                    [["NZCE Category", "tCO2e", "% of Total"]]
                    + [
                        [row.label, f"{row.value:.4f}", f"{row.pct_total:.1f}%" if row.pct_total is not None else "—"]
                        for row in report.nzce_categories
                    ],
                    [8 * cm, 3 * cm, 3 * cm],
                ),
                Spacer(1, 0.15 * cm),
                Paragraph(_pdf_text(report.nzce_note), styles["BodySmall"]),
                Spacer(1, 0.35 * cm),
            ]
        )

    if benchmark:
        story.extend(
            [
                Paragraph("Benchmark", styles["SectionHeading"]),
                Spacer(1, 0.2 * cm),
                _table(
                    [["Metric", "Value"]] + [[_labelize(key), value] for key, value in benchmark.items()],
                    [8 * cm, 6 * cm],
                ),
                Spacer(1, 0.35 * cm),
            ]
        )

    story.extend(
        [
            Paragraph("Compliance Status", styles["SectionHeading"]),
            Spacer(1, 0.2 * cm),
            _table(
                [["Framework", "Status", "Score %"]]
                + [[check.framework, check.status, f"{check.score_pct:.1f}"] for check in report.compliance.checks],
                [8 * cm, 4 * cm, 2.5 * cm],
            ),
            Spacer(1, 0.35 * cm),
            Paragraph("Offset Coverage", styles["SectionHeading"]),
            Spacer(1, 0.2 * cm),
            _table(
                [["Metric", "Value"]]
                + [[_labelize(key), value] for key, value in offset_summary.items() if not isinstance(value, dict)],
                [8 * cm, 6 * cm],
            ),
            Spacer(1, 0.35 * cm),
            Paragraph("Factor Provenance", styles["SectionHeading"]),
            Spacer(1, 0.2 * cm),
            _table(factor_rows[:9], [8 * cm, 6 * cm]),
        ]
    )

    if assumption_paragraphs:
        story.extend([Spacer(1, 0.35 * cm), Paragraph("Assumptions", styles["SectionHeading"])])
        for paragraph in assumption_paragraphs[:8]:
            story.extend([Spacer(1, 0.1 * cm), paragraph])

    # Sign-off block: signature lines for the report preparer and approver.
    sig_table = Table(
        [
            [Paragraph("<b>Prepared by</b>", styles["BodySmall"]),
             Paragraph("<b>Approved by</b>", styles["BodySmall"])],
            ["", ""],
            ["Signature", "Signature"],
            ["Name / Title", "Name / Title"],
            ["Date", "Date"],
        ],
        colWidths=[7.5 * cm, 7.5 * cm],
        rowHeights=[None, 1.2 * cm, None, None, None],
    )
    sig_table.setStyle(
        TableStyle(
            [
                ("LINEBELOW", (0, 1), (0, 1), 0.7, ink),
                ("LINEBELOW", (1, 1), (1, 1), 0.7, ink),
                ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 2), (-1, -1), 8),
                ("TEXTCOLOR", (0, 2), (-1, -1), muted),
                ("TOPPADDING", (0, 2), (-1, 2), 2),
                ("BOTTOMPADDING", (0, 2), (-1, 2), 10),
                ("BOTTOMPADDING", (0, 3), (-1, 3), 10),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 36),
                ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
            ]
        )
    )
    story.append(
        KeepTogether(
            [
                Spacer(1, 0.7 * cm),
                Paragraph("Sign-off", styles["SectionHeading"]),
                Spacer(1, 0.25 * cm),
                Paragraph(
                    "This report was generated by CutCarbon EventCarbon Co-Pilot from the scenario "
                    "inputs and emission factors recorded above.",
                    styles["BodySmall"],
                ),
                Spacer(1, 0.3 * cm),
                sig_table,
            ]
        )
    )

    doc.build(story, canvasmaker=_ReportCanvas)
    return buf.getvalue()


def _build_scenarios_workbook(rows):
    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "Scenarios"
    headers = [
        "ID", "Name", "Location", "Event Type", "Attendees", "Days", "Mode",
        "Travel tCO2e", "Venue Energy tCO2e", "Accommodation tCO2e",
        "Catering tCO2e", "Materials & Waste tCO2e",
        "Equipment tCO2e", "Swag tCO2e", "Digital tCO2e",
        "Total tCO2e", "Per Attendee tCO2e",
        "Scope 1", "Scope 2", "Scope 3",
        "Data Quality", "Created At",
    ]
    ws.append(headers)
    _style_header(ws)
    for s in rows:
        ws.append([
            s.id, s.name, _scenario_location(s),
            s.event_type, s.attendees, s.event_days, s.mode,
            round(s.travel_tco2e or 0, 4),
            round(s.venue_energy_tco2e or 0, 4),
            round(s.accommodation_tco2e or 0, 4),
            round(s.catering_tco2e or 0, 4),
            round(s.materials_waste_tco2e or 0, 4),
            round(getattr(s, "equipment_tco2e", 0) or 0, 4),
            round(getattr(s, "swag_tco2e", 0) or 0, 4),
            round(getattr(s, "digital_tco2e", 0) or 0, 4),
            round(s.total_tco2e or 0, 4),
            round(s.per_attendee_tco2e or 0, 4),
            round(s.scope1_tco2e or 0, 4),
            round(s.scope2_tco2e or 0, 4),
            round(s.scope3_tco2e or 0, 4),
            s.data_quality,
            s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws2 = wb.create_sheet("Input Payloads")
    ws2.append(["Scenario ID", "Scenario Name", "Input Payload (JSON)"])
    _style_header(ws2)
    for s in rows:
        ws2.append([s.id, s.name, json.dumps(s.input_payload or {}, indent=2)])
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 80

    return wb


@router.get("/scenarios.xlsx", summary="Download all scenarios as Excel")
async def export_scenarios_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: UserDB = Depends(get_current_user),
):
    q = (
        select(ScenarioDB)
        .where(ScenarioDB.user_id == current_user.id)
        .order_by(desc(ScenarioDB.created_at))
    )
    rows = (await db.execute(q)).scalars().all()

    content = await asyncio.to_thread(lambda: _wb_bytes(_build_scenarios_workbook(rows)))
    filename = f"cutcarbon_scenarios_{utcnow().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(content, filename)


@router.get("/scenarios/{scenario_id}.json", response_model=ScenarioReportPayload, summary="Download single scenario report package as JSON")
async def export_scenario_json(
    scenario_id: str,
    region: str = "singapore",
    has_scope3: bool = True,
    has_ghg_report: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: UserDB = Depends(get_current_user),
):
    report = await build_scenario_report_payload(
        scenario_id=scenario_id,
        db=db,
        current_user=current_user,
        region=region,
        has_scope3=has_scope3,
        has_ghg_report=has_ghg_report,
    )
    return JSONResponse(
        content=report.model_dump(),
        headers={"Content-Disposition": f'attachment; filename="{_scenario_report_filename("cutcarbon_report", scenario_id, "json")}"'},
    )


@router.get("/scenarios/{scenario_id}.csv", summary="Download single scenario report package as CSV")
async def export_scenario_csv(
    scenario_id: str,
    region: str = "singapore",
    has_scope3: bool = True,
    has_ghg_report: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: UserDB = Depends(get_current_user),
):
    report = await build_scenario_report_payload(
        scenario_id=scenario_id,
        db=db,
        current_user=current_user,
        region=region,
        has_scope3=has_scope3,
        has_ghg_report=has_ghg_report,
    )
    content = await asyncio.to_thread(_scenario_report_csv_bytes, report)
    return _csv_response(
        content,
        _scenario_report_filename("cutcarbon_report", scenario_id, "csv"),
    )


@router.get("/scenarios/{scenario_id}.xlsx", summary="Download single scenario report package as Excel")
async def export_scenario_xlsx(
    scenario_id: str,
    region: str = "singapore",
    has_scope3: bool = True,
    has_ghg_report: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: UserDB = Depends(get_current_user),
):
    report = await build_scenario_report_payload(
        scenario_id=scenario_id,
        db=db,
        current_user=current_user,
        region=region,
        has_scope3=has_scope3,
        has_ghg_report=has_ghg_report,
    )
    content = await asyncio.to_thread(lambda: _wb_bytes(_scenario_report_xlsx(report)))
    return _xlsx_response(content, _scenario_report_filename("cutcarbon_report", scenario_id, "xlsx"))


@router.get("/scenarios/{scenario_id}.pdf", summary="Download single scenario report package as PDF")
async def export_scenario_pdf(
    scenario_id: str,
    region: str = "singapore",
    has_scope3: bool = True,
    has_ghg_report: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: UserDB = Depends(get_current_user),
):
    report = await build_scenario_report_payload(
        scenario_id=scenario_id,
        db=db,
        current_user=current_user,
        region=region,
        has_scope3=has_scope3,
        has_ghg_report=has_ghg_report,
    )
    content = await asyncio.to_thread(_scenario_report_pdf, report)
    return _pdf_response(
        content,
        _scenario_report_filename("cutcarbon_report", scenario_id, "pdf"),
    )


def _build_factor_workbook():
    from openpyxl import Workbook

    ef_path = _DATA_DIR / "emission_factors.json"
    with open(ef_path, encoding="utf-8") as f:
        ef = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Emission Factors"
    ws.append(["Category", "Subcategory / Key", "Region", "Factor / Value", "Unit", "Source", "Last Fetched"])
    _style_header(ws)

    def _flatten(data, prefix=""):
        for k, v in data.items():
            if k.startswith("_") or k == "last_agent_update":
                continue
            if isinstance(v, dict):
                if "factor" in v or "economy" in v or "factor_value" in v:
                    factor = v.get("factor") or v.get("economy") or v.get("factor_value", "")
                    ws.append([
                        prefix, k,
                        v.get("region", ""),
                        factor,
                        v.get("unit", ""),
                        v.get("source", ""),
                        v.get("last_fetched", ""),
                    ])
                else:
                    _flatten(v, prefix=f"{prefix}/{k}" if prefix else k)
            elif isinstance(v, (int, float)):
                ws.append([prefix, k, "", v, "", "", ""])

    _flatten(ef)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    return wb


@router.get("/emission-factors.xlsx", summary="Download emission factor catalog as Excel")
async def export_emission_factors_xlsx(
    current_user: UserDB = Depends(get_current_user),
):
    content = await asyncio.to_thread(lambda: _wb_bytes(_build_factor_workbook()))
    filename = f"cutcarbon_emission_factors_{utcnow().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(content, filename)


def _build_agent_runs_workbook(rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Agent Runs"
    ws.append(["ID", "Agent Name", "Category", "Status", "Run ID", "Steps", "Source URL", "Fetched At", "Error", "Data (JSON)"])
    _style_header(ws)

    for r in rows:
        ws.append([
            r.id, r.agent_name, r.category, r.status,
            r.run_id or "", r.num_steps or "",
            r.source_url or "",
            r.fetched_at.strftime("%Y-%m-%d %H:%M:%S") if r.fetched_at else "",
            r.error or "",
            json.dumps(r.result_json or {}),
        ])

    col_widths = [6, 22, 16, 10, 36, 8, 55, 20, 30, 60]
    for i, width in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter

        ws.column_dimensions[get_column_letter(i)].width = width

    return wb


@router.get("/agent-runs.xlsx", summary="Download TinyFish agent run history as Excel")
async def export_agent_runs_xlsx(
    db: AsyncSession = Depends(get_db),
    limit: int = 500,
    current_user: UserDB = Depends(get_current_user),
):
    rows = (
        await db.execute(
            select(AgentRunDB).order_by(desc(AgentRunDB.fetched_at)).limit(limit)
        )
    ).scalars().all()

    content = await asyncio.to_thread(lambda: _wb_bytes(_build_agent_runs_workbook(rows)))
    filename = f"cutcarbon_agent_runs_{utcnow().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(content, filename)


def _load_factor_json() -> dict:
    ef_path = _DATA_DIR / "emission_factors.json"
    with open(ef_path, encoding="utf-8") as f:
        return json.load(f)


@router.get("/emission-factors.json", summary="Download emission_factors.json (raw)")
async def export_emission_factors_json(
    current_user: UserDB = Depends(get_current_user),
):
    data = await asyncio.to_thread(_load_factor_json)
    return JSONResponse(
        content=data,
        headers={"Content-Disposition": f'attachment; filename="emission_factors_{utcnow().strftime("%Y%m%d")}.json"'},
    )


@router.get("/scenarios.json", summary="Download all scenarios as raw JSON")
async def export_scenarios_json(
    db: AsyncSession = Depends(get_db),
    current_user: UserDB = Depends(get_current_user),
):
    q = (
        select(ScenarioDB)
        .where(ScenarioDB.user_id == current_user.id)
        .order_by(desc(ScenarioDB.created_at))
    )
    rows = (await db.execute(q)).scalars().all()

    data = [serialize_scenario(s) for s in rows]
    return JSONResponse(
        content=data,
        headers={"Content-Disposition": f'attachment; filename="cutcarbon_scenarios_{utcnow().strftime("%Y%m%d")}.json"'},
    )

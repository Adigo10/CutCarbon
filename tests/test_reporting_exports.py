import asyncio
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import select

import app.models.database as database
from app.models.database import UserDB
from app.models.schemas import EventScenarioInput, OffsetProjectType, OffsetPurchaseCreate, TravelMode, TravelSegment, VenueEnergy
from app.routers.exports import build_scenario_report_payload
from app.services.emissions_engine import calculate_scenario

from helpers import create_seeded_scenario, register_user


def load_user(email: str) -> UserDB:
    async def _runner():
        async with database.AsyncSessionLocal() as session:
            result = await session.execute(select(UserDB).where(UserDB.email == email))
            return result.scalar_one()

    return asyncio.run(_runner())


def build_payload_for_test(scenario_id: str, user: UserDB):
    async def _runner():
        async with database.AsyncSessionLocal() as session:
            return await build_scenario_report_payload(
                scenario_id=scenario_id,
                db=session,
                current_user=user,
                region="eu",
                has_scope3=False,
                has_ghg_report=False,
            )

    return asyncio.run(_runner())


def test_shared_report_payload_includes_offsets_and_compliance_overrides(client: TestClient):
    headers = register_user(client)
    scenario_id = create_seeded_scenario(client, headers)
    user = load_user("reporter@example.com")

    report = build_payload_for_test(scenario_id, user)

    assert report.scenario["scenario_id"] == scenario_id
    assert report.offset_portfolio.total_retired_tco2e == pytest.approx(2.5)
    assert report.offset_portfolio.coverage_pct is not None
    assert report.compliance_overrides.region == "eu"
    assert report.compliance_overrides.has_scope3 is False
    assert "EU CSRD" in report.compliance.mandatory_frameworks
    assert report.categories

    # NZCE mapping: all 9 methodology categories present, totals conserved.
    assert len(report.nzce_categories) == 9
    nzce_total = sum(row.value for row in report.nzce_categories)
    assert nzce_total == pytest.approx(report.scenario["emissions"]["total_tco2e"], abs=0.01)
    labels = [row.label for row in report.nzce_categories]
    assert "Digital Content & Communication" in labels
    assert "Local Transportation" in labels
    assert report.nzce_note

    # NZCE also appears as a compliance framework check.
    assert any("Net Zero Carbon Events" in c.framework for c in report.compliance.checks)

    # Per-category data-quality disclosure persisted via assumptions.
    quality = report.assumptions["category_data_quality"]
    assert quality["travel"] == "actual"
    assert quality["waste"] == "proxy"


@pytest.mark.parametrize("fmt", ["json", "csv", "xlsx", "pdf"])
def test_single_scenario_report_exports_require_auth(client: TestClient, fmt: str):
    response = client.get(f"/api/exports/scenarios/missing.{fmt}")
    assert response.status_code == 401


@pytest.mark.parametrize("fmt", ["json", "csv", "xlsx", "pdf"])
def test_single_scenario_report_exports_return_expected_files(client: TestClient, fmt: str):
    headers = register_user(client, email=f"{fmt}@example.com")
    scenario_id = create_seeded_scenario(client, headers)

    response = client.get(
        f"/api/exports/scenarios/{scenario_id}.{fmt}"
        "?region=eu&has_scope3=false&has_ghg_report=false",
        headers=headers,
    )

    assert response.status_code == 200
    assert f".{fmt}" in response.headers["content-disposition"]

    if fmt == "json":
        payload = response.json()
        assert payload["compliance_overrides"]["region"] == "eu"
        assert payload["compliance_overrides"]["has_scope3"] is False
        assert "EU CSRD" in payload["compliance"]["mandatory_frameworks"]
        assert len(payload["nzce_categories"]) == 9
    elif fmt == "csv":
        content = response.content.decode("utf-8")
        assert "section,key,label,value,unit" in content
        assert "metadata,region,Compliance Region,eu," in content
        assert "compliance,overall_score_pct,Overall Score" in content
        assert "nzce,nzce_energy,Energy" in content
    elif fmt == "xlsx":
        workbook = load_workbook(io.BytesIO(response.content))
        assert "Report Summary" in workbook.sheetnames
        assert "Compliance" in workbook.sheetnames
        assert "NZCE Mapping" in workbook.sheetnames
    else:
        assert response.content.startswith(b"%PDF")


def test_global_exports_still_work(client: TestClient):
    headers = register_user(client, email="global@example.com")

    # Factor-catalog exports now REQUIRE auth (previously an unauthenticated gap).
    assert client.get("/api/exports/emission-factors.json").status_code == 401
    assert client.get("/api/exports/emission-factors.xlsx").status_code == 401

    factor_json = client.get("/api/exports/emission-factors.json", headers=headers)
    assert factor_json.status_code == 200
    assert "travel" in factor_json.json()

    factor_xlsx = client.get("/api/exports/emission-factors.xlsx", headers=headers)
    assert factor_xlsx.status_code == 200
    assert ".xlsx" in factor_xlsx.headers["content-disposition"]

    agent_xlsx = client.get("/api/exports/agent-runs.xlsx", headers=headers)
    assert agent_xlsx.status_code == 200
    assert ".xlsx" in agent_xlsx.headers["content-disposition"]


def test_invalid_emissions_inputs_are_rejected():
    with pytest.raises(ValidationError):
        EventScenarioInput(
            name="Invalid Travel",
            attendees=100,
            travel_segments=[
                TravelSegment(
                    mode=TravelMode.LONG_HAUL_FLIGHT,
                    attendees=-50,
                    distance_km=2000,
                )
            ],
        )

    with pytest.raises(ValidationError):
        EventScenarioInput(
            name="Invalid Renewable",
            attendees=100,
            venue_energy=VenueEnergy(kwh_consumed=1000, renewable_pct=150),
        )

    with pytest.raises(ValidationError):
        EventScenarioInput(
            name="Overallocated Travel",
            attendees=100,
            travel_segments=[
                TravelSegment(
                    mode=TravelMode.LONG_HAUL_FLIGHT,
                    attendees=101,
                    distance_km=2000,
                )
            ],
        )

    with pytest.raises(ValidationError):
        OffsetPurchaseCreate(
            project_type=OffsetProjectType.RENEWABLE_ENERGY,
            quantity_tco2e=-5,
            price_per_tco2e_usd=10,
        )


def test_zero_kwh_is_treated_as_actual_data():
    scenario = EventScenarioInput(
        name="Zero kWh",
        attendees=10,
        venue_energy=VenueEnergy(kwh_consumed=0, venue_area_m2=100),
    )

    result = calculate_scenario(scenario)

    assert result.emissions.venue_energy_tco2e == 0
    assert result.assumptions["venue"] == "Actual kWh: 0"


def test_offset_purchase_requires_owned_scenario(client: TestClient):
    headers = register_user(client, email="offset-owner@example.com")

    response = client.post(
        "/api/offsets",
        json={
            "scenario_id": "missing-scenario",
            "project_type": "renewable_energy",
            "registry": "gold_standard",
            "quantity_tco2e": 1,
            "price_per_tco2e_usd": 10,
            "vintage_year": 2025,
        },
        headers=headers,
    )

    assert response.status_code == 404


# --- Security & de-gimmick regressions -------------------------------------------------

# Forged/bad-signature token rejection is covered by tests/test_auth.py
# (test_bad_signature_rejected, test_wrong_issuer_rejected, ...) under the Supabase
# JWKS verifier — the old PLACEHOLDER_JWT_SECRET forgery test was removed with that auth model.


def test_csv_export_neutralizes_formula_injection(client: TestClient):
    headers = register_user(client, email="csvinj@example.com")
    created = client.post(
        "/api/scenarios",
        json={"name": "=1+1", "attendees": 50, "event_days": 1, "location": "Singapore"},
        headers=headers,
    )
    assert created.status_code == 200
    scenario_id = created.json()["scenario_id"]

    csv_resp = client.get(f"/api/exports/scenarios/{scenario_id}.csv", headers=headers)
    assert csv_resp.status_code == 200
    body = csv_resp.text
    # The dangerous name is stored as literal text (prefixed with a quote), never a bare formula.
    assert "'=1+1" in body


def test_financial_total_excludes_fabricated_streams():
    from app.models.schemas import FinancialRequest
    from app.services.financial_engine import generate_financial_report

    req = FinancialRequest(
        baseline_tco2e=100, reduced_tco2e=70, region="singapore",
        energy_kwh_saved=1000, meal_switches=200, attendees=300,
        actions_taken=["ghg_reporting", "renewable_energy"],
    )
    res = generate_financial_report(req)

    schemes = [t.scheme for t in res.carbon_tax_savings]
    assert not any("Voluntary Carbon Market" in s for s in schemes)  # VCM cost not sold as a saving
    assert res.compliance_value_usd == 0.0                            # no $50k/$5k magic
    assert res.roi_months is None                                     # no one-time/12 ROI
    primary = res.carbon_tax_savings[0].savings_usd if res.carbon_tax_savings else 0
    expected = round(primary + res.energy_cost_savings_usd + res.catering_cost_savings_usd, 2)
    assert abs(res.total_financial_savings_usd - expected) < 0.01


def test_financial_uses_live_carbon_price_when_present():
    from app.services import emissions_engine as eng
    from app.services.financial_engine import calculate_carbon_tax_savings

    live = eng.EF.setdefault("carbon_tax_live", {})
    live["singapore_current_sgd"] = 99
    try:
        savings = calculate_carbon_tax_savings(10, "singapore")
        assert savings[0].savings_local == 990  # 10 t x 99 SGD
        assert "live price" in savings[0].description
    finally:
        live.pop("singapore_current_sgd", None)


def test_reduction_suggestions_are_clamped_and_offsets_separated():
    from app.models.schemas import CateringGroup
    from app.services.emissions_engine import calculate_scenario, get_reduction_suggestions

    scenario = EventScenarioInput(
        name="Vegan dinner", attendees=500, event_days=2,
        catering=CateringGroup(catering_type="vegan_meal", meals=2000),
    )
    result = calculate_scenario(scenario)
    suggestions = get_reduction_suggestions(result, 30.0, catering_type="vegan_meal")

    actions = [s["action"] for s in suggestions]
    assert "vegetarian_menu" not in actions  # already plant-based -> gated off

    reductions = [s for s in suggestions if not s.get("is_neutralization")]
    assert sum(s["co2e_saved_tco2e"] for s in reductions) <= result.emissions.total_tco2e + 1e-6
    # Offsets are a separate neutralization entry, listed last (never ranked as a reduction).
    offsets = [s for s in suggestions if s.get("is_neutralization")]
    assert offsets and suggestions[-1].get("is_neutralization")


def test_tinyfish_extract_validates_and_converts():
    from app.services.tinyfish_agent import (
        EUGridFactorAgent,
        FlightEmissionFactorAgent,
        SingaporeGridFactorAgent,
    )

    assert SingaporeGridFactorAgent().extract({"factor": 0.402, "unit": "kg_co2e_per_kwh"})["factor_value"] == 0.402
    assert EUGridFactorAgent().extract({"factor": 213, "unit": "g_co2e_per_kwh"})["factor_value"] == 0.213
    # A year mis-parsed as the factor is out of bounds -> rejected (keeps prior good value).
    assert SingaporeGridFactorAgent().extract({"factor": 2024})["factor_value"] is None
    # Business cheaper than economy is implausible -> dropped.
    flight = FlightEmissionFactorAgent().extract(
        {"short_haul_economy": 0.15, "long_haul_economy": 0.19, "long_haul_business": 0.10}
    )
    assert flight["long_haul_business"] is None
